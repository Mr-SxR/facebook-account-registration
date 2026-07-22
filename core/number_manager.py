import os
import re
import threading
import openpyxl
from ui.colors import GREEN, RED, WHITE, EKL, LINE
from core.settings_manager import load_settings

file_lock = threading.Lock()

MAX_FORGET_NUMBERS = 10000
MAX_CONFIRM_ENTRIES = 5000
CONFIRM_FILE = "Confirm_List.txt"


# Extract phone numbers from the best-matching column in an Excel file
def extract_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        sheet = wb.active
        target_col = None
        max_matches = 0

        # Scan first 20 rows to find the column with the most phone-like values
        for col in range(1, sheet.max_column + 1):
            matches = 0
            for row in range(2, min(22, sheet.max_row + 1)):
                val = sheet.cell(row=row, column=col).value
                if val:
                    cleaned = re.sub(r'[\s\-\(\)\+]', '', str(val).strip())
                    if cleaned.isdigit() and 7 <= len(cleaned) <= 15:
                        matches += 1
            if matches > max_matches:
                max_matches = matches
                target_col = col

        if target_col is None:
            return None, "No phone number column found."

        numbers = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=target_col, max_col=target_col, values_only=True):
            val = row[0]
            if val:
                cleaned = re.sub(r'[\s\-\(\)\+]', '', str(val).strip())
                if cleaned.isdigit() and 7 <= len(cleaned) <= 15:
                    numbers.append(cleaned)
        return numbers, None
    except Exception as e:
        return None, str(e)


# Load phone numbers from Number_List.txt
def load_numbers():
    with file_lock:
        try:
            with open("Number_List.txt", "r", encoding="utf-8", errors="ignore") as f:
                return [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            return []
        except Exception:
            return []


# Save phone numbers to Number_List.txt
def save_numbers(numbers):
    with file_lock:
        with open("Number_List.txt", "w", encoding="utf-8") as f:
            for n in numbers:
                f.write(n + "\n")


# Remove a processed number from Number_List.txt (thread-safe)
def remove_number(number):
    with file_lock:
        try:
            with open("Number_List.txt", "r", encoding="utf-8", errors="ignore") as f:
                lines = [line.strip() for line in f if line.strip()]
        except (FileNotFoundError, Exception):
            return

        if number in lines:
            lines.remove(number)

        try:
            with open("Number_List.txt", "w", encoding="utf-8") as f:
                for n in lines:
                    f.write(n + "\n")
        except Exception:
            pass


# Route file input based on settings (txt, multi-excel, or auto-detect)
def process_file_input():
    settings = load_settings()
    file_cfg = settings.get("file_input_settings", {})
    always_txt = file_cfg.get("always_use_txt", False)
    multi_excel = file_cfg.get("use_multiple_excel_files", False)

    if always_txt:
        return _load_txt()
    if multi_excel:
        return _load_multi_excel()
    return _load_auto()


# Load numbers directly from Number_List.txt
def _load_txt():
    if not os.path.exists("Number_List.txt"):
        print(f"{WHITE} 'Number_List.txt' file was not found.")
        return []
    numbers = load_numbers()
    if not numbers:
        print(f"{WHITE} 'Number_List.txt' file is empty.")
        return []
    if len(numbers) > MAX_FORGET_NUMBERS:
        print(f"{RED} Too many numbers! Maximum {MAX_FORGET_NUMBERS} allowed.")
        print(f"{RED} You have {len(numbers)} numbers. Please reduce and try again.")
        return None
    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} Number_List.txt")
    return numbers


# Load and merge numbers from all Excel files in the directory
def _load_multi_excel():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_txt()

    print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files.")
    all_numbers = []
    for f in xlsx_files:
        print(f"{WHITE} Extracting from {EKL} {f}...")
        nums, err = extract_from_excel(f)
        if nums:
            all_numbers.extend(nums)
            print(f"{GREEN}  -> Found {len(nums)} numbers.")
        else:
            print(f"{RED}  -> Failed: {err}")

    if not all_numbers:
        print(f"{RED} No valid numbers found in any Excel files.")
        return None

    all_numbers = list(set(all_numbers))

    if len(all_numbers) > MAX_FORGET_NUMBERS:
        print(f"{RED} Too many numbers! Maximum {MAX_FORGET_NUMBERS} allowed.")
        print(f"{RED} Total found: {len(all_numbers)}. Please reduce files and try again.")
        return None

    save_numbers(all_numbers)
    print(f"\n {GREEN}[{RED}●{GREEN}] Total Unique Numbers {EKL} {len(all_numbers)}")
    print(f" {GREEN}[{RED}●{GREEN}] Saved to 'Number_List.txt'\n")
    return all_numbers


# Auto-detect input source: use Excel if available, fall back to txt
def _load_auto():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_txt()

    filename = None
    if len(xlsx_files) == 1:
        filename = xlsx_files[0]
    else:
        print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files:")
        for idx, f in enumerate(xlsx_files, 1):
            print(f" {GREEN}[{RED}{idx}{GREEN}] {f}")
        print(f"{LINE}")
        while True:
            try:
                choice = input(f" {GREEN}[{RED}●{GREEN}] Select File (1-{len(xlsx_files)}) {EKL} ").strip()
                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(xlsx_files):
                        filename = xlsx_files[idx]
                        break
                print(f"{RED} Invalid selection!")
            except KeyboardInterrupt:
                raise
            except Exception:
                pass

    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} {filename}\n")
    nums, err = extract_from_excel(filename)

    if nums:
        if len(nums) > MAX_FORGET_NUMBERS:
            print(f"{RED} Too many numbers! Maximum {MAX_FORGET_NUMBERS} allowed.")
            print(f"{RED} Found {len(nums)} numbers in {filename}. Please reduce and try again.")
            return None
        save_numbers(nums)
        print(f" {GREEN}[{RED}●{GREEN}] Extracted {len(nums)} numbers from {filename}")
        print(f" {GREEN}[{RED}●{GREEN}] Saved to 'Number_List.txt'\n")
        return nums
    else:
        print(f"{RED} Error: {err}")
        return None


# Load confirm list entries (number|otp format) from Confirm_List.txt
def load_confirm_list():
    with file_lock:
        try:
            with open(CONFIRM_FILE, "r", encoding="utf-8", errors="ignore") as f:
                entries = []
                for line in f:
                    line = line.strip()
                    if not line or "|" not in line:
                        continue
                    parts = line.split("|", 1)
                    number = parts[0].strip()
                    otp = parts[1].strip()
                    if number and otp:
                        entries.append((number, otp))
                if not entries:
                    print(f"{WHITE} '{CONFIRM_FILE}' is empty or has no valid entries.")
                    return []
                if len(entries) > MAX_CONFIRM_ENTRIES:
                    print(f"{RED} Too many entries! Maximum {MAX_CONFIRM_ENTRIES} allowed.")
                    print(f"{RED} You have {len(entries)} entries. Please reduce and try again.")
                    return []
                return entries
        except FileNotFoundError:
            print(f"{WHITE} '{CONFIRM_FILE}' file not found.")
            return []
        except Exception:
            print(f"{RED} Error reading '{CONFIRM_FILE}'.")
            return []


FILTER_INPUT_FILE = "Filter_Input_Num.txt"
MAX_FILTER_NUMBERS = 50000


# Route filter input based on settings
def load_filter_input():
    settings = load_settings()
    file_cfg = settings.get("file_input_settings", {})
    always_txt = file_cfg.get("always_use_txt", False)
    multi_excel = file_cfg.get("use_multiple_excel_files", False)

    if always_txt:
        return _load_filter_txt()
    if multi_excel:
        return _load_filter_multi_excel()
    return _load_filter_auto()


# Load filter numbers from Filter_Input_Num.txt
def _load_filter_txt():
    if not os.path.exists(FILTER_INPUT_FILE):
        print(f"{WHITE} '{FILTER_INPUT_FILE}' file was not found.")
        return []
    numbers = load_filter_numbers()
    if not numbers:
        print(f"{WHITE} '{FILTER_INPUT_FILE}' file is empty.")
        return []
    if len(numbers) > MAX_FILTER_NUMBERS:
        print(f"{RED} Too many numbers! Maximum {MAX_FILTER_NUMBERS} allowed.")
        print(f"{RED} You have {len(numbers)} numbers. Please reduce and try again.")
        return None
    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} {FILTER_INPUT_FILE}")
    return numbers


# Read raw filter numbers from file
def load_filter_numbers():
    with file_lock:
        try:
            with open(FILTER_INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
                return [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            return []
        except Exception:
            return []


# Save filter numbers to Filter_Input_Num.txt
def save_filter_numbers(numbers):
    with file_lock:
        with open(FILTER_INPUT_FILE, "w", encoding="utf-8") as f:
            for n in numbers:
                f.write(n + "\n")


# Load and merge filter numbers from all Excel files
def _load_filter_multi_excel():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_filter_txt()

    print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files.")
    all_numbers = []
    for f in xlsx_files:
        print(f"{WHITE} Extracting from {EKL} {f}...")
        nums, err = extract_from_excel(f)
        if nums:
            all_numbers.extend(nums)
            print(f"{GREEN}  -> Found {len(nums)} numbers.")
        else:
            print(f"{RED}  -> Failed: {err}")

    if not all_numbers:
        print(f"{RED} No valid numbers found in any Excel files.")
        return None

    all_numbers = list(set(all_numbers))

    if len(all_numbers) > MAX_FILTER_NUMBERS:
        print(f"{RED} Too many numbers! Maximum {MAX_FILTER_NUMBERS} allowed.")
        print(f"{RED} Total found: {len(all_numbers)}. Please reduce files and try again.")
        return None

    save_filter_numbers(all_numbers)
    print(f"\n {GREEN}[{RED}●{GREEN}] Total Unique Numbers {EKL} {len(all_numbers)}")
    print(f" {GREEN}[{RED}●{GREEN}] Saved to '{FILTER_INPUT_FILE}'\n")
    return all_numbers


# Auto-detect filter input source
def _load_filter_auto():
    xlsx_files = [f for f in os.listdir('.') if f.endswith(".xlsx") and not f.startswith("~$")]
    if not xlsx_files:
        return _load_filter_txt()

    filename = None
    if len(xlsx_files) == 1:
        filename = xlsx_files[0]
    else:
        print(f" {GREEN}[{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files:")
        for idx, f in enumerate(xlsx_files, 1):
            print(f" {GREEN}[{RED}{idx}{GREEN}] {f}")
        print(f"{LINE}")
        while True:
            try:
                choice = input(f" {GREEN}[{RED}●{GREEN}] Select File (1-{len(xlsx_files)}) {EKL} ").strip()
                if choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(xlsx_files):
                        filename = xlsx_files[idx]
                        break
                print(f"{RED} Invalid selection!")
            except KeyboardInterrupt:
                raise
            except Exception:
                pass

    print(f" {GREEN}[{RED}●{GREEN}] Selected File {EKL} {filename}\n")
    nums, err = extract_from_excel(filename)

    if nums:
        if len(nums) > MAX_FILTER_NUMBERS:
            print(f"{RED} Too many numbers! Maximum {MAX_FILTER_NUMBERS} allowed.")
            print(f"{RED} Found {len(nums)} numbers in {filename}. Please reduce and try again.")
            return None
        save_filter_numbers(nums)
        print(f" {GREEN}[{RED}●{GREEN}] Extracted {len(nums)} numbers from {filename}")
        print(f" {GREEN}[{RED}●{GREEN}] Saved to '{FILTER_INPUT_FILE}'\n")
        return nums
    else:
        print(f"{RED} Error: {err}")
        return None


# Remove a processed number from Filter_Input_Num.txt (thread-safe)
def remove_filter_number(number):
    with file_lock:
        try:
            with open(FILTER_INPUT_FILE, "r", encoding="utf-8", errors="ignore") as f:
                lines = [line.strip() for line in f if line.strip()]
        except (FileNotFoundError, Exception):
            return
        if number in lines:
            lines.remove(number)
        try:
            with open(FILTER_INPUT_FILE, "w", encoding="utf-8") as f:
                for n in lines:
                    f.write(n + "\n")
        except Exception:
            pass


# Remove a specific confirm entry (number|otp pair) from Confirm_List.txt
def remove_confirm_entry(number, otp):
    with file_lock:
        try:
            with open(CONFIRM_FILE, "r", encoding="utf-8", errors="ignore") as f:
                lines = [line.strip() for line in f if line.strip()]
        except (FileNotFoundError, Exception):
            return

        new_lines = []
        for line in lines:
            if "|" in line:
                parts = line.split("|", 1)
                if parts[0].strip() == number and parts[1].strip() == otp:
                    continue
            new_lines.append(line)

        try:
            with open(CONFIRM_FILE, "w", encoding="utf-8") as f:
                for line in new_lines:
                    f.write(line + "\n")
        except Exception:
            pass
